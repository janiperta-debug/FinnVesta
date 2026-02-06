"""
Building Import API

Handles bulk import of buildings from Excel/CSV files.
Supports smart column mapping with auto-detection of Finnish column names.
"""

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import List, Dict, Optional, Any
from enum import Enum
import openpyxl
import pandas as pd
import io
import asyncpg
import os
from app.auth import AuthorizedUser
import uuid
import databutton as db
import difflib
from datetime import datetime

router = APIRouter(prefix="/buildings/import")


# In-memory storage for uploaded files during import session
# Key: file_id, Value: (file_content, filename, header_row)
UPLOAD_CACHE: Dict[str, tuple[bytes, str, int]] = {}


async def get_db_connection():
    """Get asyncpg connection to database"""
    return await asyncpg.connect(os.environ.get("DATABASE_URL"))


# Column mapping patterns for auto-detection
COLUMN_MAPPINGS = {
    'name': ['kohde', 'rakennus', 'nimi', 'building', 'name'],
    'address': ['katuosoite', 'osoite', 'address', 'sijainti', 'location'],
    'construction_year': ['rakennusvuosi', 'vuosi', 'built', 'year', 'valmistumisvuosi'],
    'area_m2': ['bruttoala', 'pinta-ala', 'pinta ala', 'area', 'm2', 'neliöt', 'koko', 'bruttopinta-ala'],
    'building_type': ['käyttötarkoitus', 'tyyppi', 'type', 'käyttö'],
    'municipality': ['kunta', 'municipality', 'kaupunki'],
}

PTS_MAPPINGS = {
    'building_name': ['rakennus', 'kohde', 'nimi', 'building'],
    'component': ['osa', 'komponentti', 'rakenneosa', 'component'],
    'year': ['vuosi', 'ajankohta', 'year', 'date'],
    'cost': ['kustannus', 'hinta', 'budjetti', 'cost', 'price', 'euro'],
    'priority': ['kiireellisyys', 'prioriteetti', 'priority', 'luokka'],
}

ASSESSMENT_MAPPINGS = {
    'building_name': ['rakennus', 'kohde', 'nimi'],
    'component': ['osa', 'komponentti', 'rakenneosa'],
    'score': ['kunto', 'arvosana', 'luokka', 'score', 'condition'],
    'date': ['pvm', 'päivämäärä', 'date', 'tarkastus'],
}

MAINTENANCE_MAPPINGS = {
    'building_name': ['rakennus', 'kohde', 'nimi'],
    'description': ['kuvaus', 'työ', 'selite', 'description', 'toimenpide'],
    'date': ['pvm', 'päivämäärä', 'date', 'valmistunut', 'aika'],
    'cost': ['kustannus', 'hinta', 'toteutunut', 'cost'],
    'component': ['osa', 'komponentti', 'rakenneosa'],
}


def parse_date(value: Any) -> Optional[datetime.date]:
    """Parse date from various formats"""
    if not value or pd.isna(value):
        return None
    
    # If it's already a datetime/timestamp
    if hasattr(value, 'date'):
        return value.date()
        
    s = str(value).strip()
    
    # Try common formats
    formats = [
        '%d.%m.%Y', # 31.12.2023
        '%Y-%m-%d', # 2023-12-31
        '%d/%m/%Y', # 31/12/2023
        '%Y.%m.%d', # 2023.12.31
        '%d.%m.%y'  # 31.12.23
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            # Fix 2-digit years if needed (Python handles %y, but let's be safe for historical data)
            return dt.date()
        except ValueError:
            continue
            
    return None

def parse_float(value: Any) -> float:
    """Parse float from string like '1 234,56 €'"""
    if not value or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    
    s = str(value).replace('€', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except:
        return 0.0


def detect_column_mapping(headers: List[str], mapping_dict: Dict[str, List[str]] = COLUMN_MAPPINGS) -> Dict[str, Optional[str]]:
    """
    Auto-detect column mappings based on header names.
    Returns dict mapping FinnVesta fields to detected column names.
    """
    headers_lower = [str(h).lower().strip() if h else '' for h in headers]
    mapping = {}
    used_headers = set()  # Track which headers have been mapped to avoid duplicates
    
    for finn_field, possible_names in mapping_dict.items():
        detected = None
        for header_idx, header in enumerate(headers_lower):
            # Skip if this header is already used
            if headers[header_idx] in used_headers:
                continue
            for possible in possible_names:
                # Exact match or contains
                if header == possible or possible in header:
                    detected = headers[header_idx]  # Use original case
                    break
            if detected:
                break
        if detected:
            used_headers.add(detected)
        mapping[finn_field] = detected
    
    return mapping


def find_header_row(file_content: bytes, filename: str) -> int:
    """
    Auto-detect which row contains the actual column headers.
    Skips title rows, date rows, and empty rows.
    
    Returns the row index (0-based) to use as header.
    """
    if filename.lower().endswith('.csv'):
        # CSV files typically have headers in first row
        return 0
    
    # For Excel files, try to find the header row
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet = wb.active
        
        # Look for the first row with multiple non-empty text cells
        for row_idx in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
            row_cells = []
            for col_idx in range(1, min(20, sheet.max_column + 1)):  # Check first 20 columns
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                    row_cells.append(cell.value.strip())
            
            # If we found a row with 3+ text values, likely the header
            if len(row_cells) >= 3:
                # Check if these look like column names (not just a title or date)
                first_val = row_cells[0].lower()
                # Skip if it's just a title or date in first cell
                if any(keyword in first_val for keyword in ['tunnus', 'kohde', 'rakennus', 'nimi', 'id', 'name', 'building']):
                    return row_idx - 1  # Convert to 0-based index
        
        # Default to row 0 if not found
        return 0
    except Exception as e:
        print(f"Error detecting header row: {e}")
        return 0


class ParseFileResponse(BaseModel):
    """Response from parsing uploaded file"""
    file_id: str  # ID to reference this file in execute step
    headers: List[str]
    preview_rows: List[Dict[str, Any]]
    suggested_mapping: Dict[str, Optional[str]]
    total_rows: int


class ColumnMapping(BaseModel):
    """Column mapping from user's file to FinnVesta fields"""
    name: Optional[str] = None
    address: Optional[str] = None
    construction_year: Optional[str] = None
    area_m2: Optional[str] = None
    building_type: Optional[str] = None
    municipality: Optional[str] = None

class PTSColumnMapping(BaseModel):
    building_name: Optional[str] = None
    component: Optional[str] = None
    year: Optional[str] = None
    cost: Optional[str] = None
    priority: Optional[str] = None

class AssessmentColumnMapping(BaseModel):
    building_name: Optional[str] = None
    component: Optional[str] = None
    score: Optional[str] = None
    date: Optional[str] = None

class MaintenanceColumnMapping(BaseModel):
    building_name: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    cost: Optional[str] = None
    component: Optional[str] = None

class ExecuteImportRequest(BaseModel):
    """Request to execute import with column mappings"""
    file_id: str  # Reference to uploaded file from parse step
    column_mapping: ColumnMapping
    skip_invalid: bool = True  # Skip invalid rows or fail entire import
    selected_rows: Optional[List[int]] = None # Indices of rows to import (0-based from preview)
    detect_sub_buildings: bool = False # Whether to auto-detect sub-buildings

class ExecutePTSImportRequest(BaseModel):
    file_id: str
    column_mapping: PTSColumnMapping
    skip_invalid: bool = True

class ExecuteAssessmentImportRequest(BaseModel):
    file_id: str
    column_mapping: AssessmentColumnMapping
    skip_invalid: bool = True

class ExecuteMaintenanceImportRequest(BaseModel):
    file_id: str
    column_mapping: MaintenanceColumnMapping
    skip_invalid: bool = True


class ImportError(BaseModel):
    """Error in a specific row"""
    row_number: int
    errors: List[str]


class ExecuteImportResponse(BaseModel):
    """Response from executing import"""
    success_count: int
    skipped_count: int
    errors: List[ImportError]
    message: str


@router.post("/parse")
async def parse_import_file(
    user: AuthorizedUser,
    file: UploadFile = File(...)
) -> ParseFileResponse:
    """
    Parse uploaded Excel or CSV file and return headers, preview, and suggested column mapping.
    
    Supports:
    - Excel (.xlsx, .xls)
    - CSV (.csv)
    
    Returns first 5 rows as preview and auto-detected column mappings.
    Stores file for later use in execute step.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Tiedoston nimi puuttuu")
    
    filename_lower = file.filename.lower()
    
    try:
        # Read file content
        content = await file.read()
        
        # Auto-detect header row
        header_row = find_header_row(content, file.filename)
        print(f"Detected header row: {header_row}")
        
        # Parse based on file type
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), header=header_row)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=header_row)
        else:
            raise HTTPException(
                status_code=400, 
                detail="Tiedostomuotoa ei tueta. Käytä .xlsx, .xls tai .csv tiedostoa."
            )
        
        # Clean up: remove completely empty rows
        df = df.dropna(how='all')
        
        # Get headers
        headers = df.columns.tolist()
        print(f"Detected columns: {headers[:10]}")
        
        # Get preview (limit to 1000 rows for selection UI)
        preview_df = df.head(1000).fillna("")
        preview_rows = preview_df.to_dict('records')
        
        # Auto-detect column mapping
        suggested_mapping = detect_column_mapping(headers, COLUMN_MAPPINGS)
        print(f"Suggested mapping: {suggested_mapping}")
        
        # Generate file ID and store for later
        file_id = str(uuid.uuid4())
        UPLOAD_CACHE[file_id] = (content, file.filename, header_row)
        
        return ParseFileResponse(
            file_id=file_id,
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=suggested_mapping,
            total_rows=len(df)
        )
        
    except Exception as e:
        print(f"Error parsing file: {e}")
        raise HTTPException(
            status_code=400,
            detail=f"Virhe tiedoston lukemisessa: {str(e)}"
        )

@router.post("/pts/parse")
async def parse_pts_file(user: AuthorizedUser, file: UploadFile = File(...)) -> ParseFileResponse:
    return await _parse_generic_file(file, PTS_MAPPINGS)

@router.post("/assessments/parse")
async def parse_assessment_file(user: AuthorizedUser, file: UploadFile = File(...)) -> ParseFileResponse:
    return await _parse_generic_file(file, ASSESSMENT_MAPPINGS)

@router.post("/maintenance/parse")
async def parse_maintenance_file(user: AuthorizedUser, file: UploadFile = File(...)) -> ParseFileResponse:
    return await _parse_generic_file(file, MAINTENANCE_MAPPINGS)

async def _parse_generic_file(file: UploadFile, mapping_dict: Dict) -> ParseFileResponse:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Tiedoston nimi puuttuu")
    
    try:
        content = await file.read()
        header_row = find_header_row(content, file.filename)
        
        filename_lower = file.filename.lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), header=header_row)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=header_row)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
            
        df = df.dropna(how='all')
        headers = df.columns.tolist()
        preview_rows = df.head(1000).fillna("").to_dict('records')
        suggested_mapping = detect_column_mapping(headers, mapping_dict)
        
        file_id = str(uuid.uuid4())
        UPLOAD_CACHE[file_id] = (content, file.filename, header_row)
        
        return ParseFileResponse(
            file_id=file_id,
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=suggested_mapping,
            total_rows=len(df)
        )
    except Exception as e:
        print(f"Error parsing file: {e}")
        raise HTTPException(status_code=400, detail=f"Virhe: {str(e)}")


async def find_building_id(conn, name: str, org_id: int) -> Optional[int]:
    """Find building ID by name with fuzzy matching"""
    if not name:
        return None
        
    # 1. Exact match
    exact = await conn.fetchval(
        "SELECT id FROM buildings WHERE org_id = $1 AND LOWER(name) = LOWER($2)",
        org_id, name
    )
    if exact:
        return exact
        
    # 2. Fetch all building names for fuzzy match
    buildings = await conn.fetch(
        "SELECT id, name FROM buildings WHERE org_id = $1",
        org_id
    )
    
    if not buildings:
        return None
        
    choices = {b['name']: b['id'] for b in buildings}
    best_match = difflib.get_close_matches(name, list(choices.keys()), n=1, cutoff=0.85)
    
    if best_match:
        return choices[best_match[0]]
        
    return None

def is_sub_building(name: str) -> bool:
    keywords = ['sauna', 'varasto', 'lato', 'vaja', 'talousrakennus', 'autotalli', 'liiteri', 'jätekatos']
    name_lower = name.lower()
    return any(k in name_lower for k in keywords)


def validate_building_row(row: Dict[str, Any], row_idx: int) -> List[str]:
    """Validate a single building row"""
    errors = []
    
    if not row.get('name'):
        errors.append("Nimi puuttuu")
        
    if not row.get('construction_year'):
        errors.append("Rakennusvuosi puuttuu")
    else:
        try:
            year = int(row['construction_year'])
            if year < 1700 or year > 2100:
                errors.append(f"Epäkelpo rakennusvuosi: {year}")
        except:
            errors.append("Rakennusvuosi ei ole numero")
            
    if not row.get('area_m2'):
        errors.append("Pinta-ala puuttuu")
    else:
        try:
            area = float(row['area_m2'])
            if area <= 0:
                errors.append("Pinta-alan on oltava positiivinen")
        except:
            errors.append("Pinta-ala ei ole numero")
            
    return errors


@router.post("/execute")
async def execute_import(
    user: AuthorizedUser,
    request: ExecuteImportRequest
) -> ExecuteImportResponse:
    """
    Execute building import with provided column mappings.
    
    Validates all rows, then batch inserts valid buildings.
    Can skip invalid rows or fail entire import based on skip_invalid flag.
    """
    
    # Retrieve stored file
    if request.file_id not in UPLOAD_CACHE:
        raise HTTPException(
            status_code=400,
            detail="Tiedostoa ei löytynyt. Lataa tiedosto uudelleen."
        )
    
    content, filename, header_row = UPLOAD_CACHE[request.file_id]
    
    try:
        # Parse the full file again
        filename_lower = filename.lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), header=header_row)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=header_row)
        else:
            raise HTTPException(status_code=400, detail="Tiedostomuoto ei kelpaa")
        
        # Clean up
        df = df.dropna(how='all')
        all_rows = df.to_dict('records')
        
        # Clean up cache after reading
        del UPLOAD_CACHE[request.file_id]
        
    except Exception as e:
        print(f"Error re-parsing file: {e}")
        raise HTTPException(
            status_code=400,
            detail=f"Virhe tiedoston lukemisessa: {str(e)}"
        )
    
    # Map columns to building fields
    mapped_rows = []
    errors: List[ImportError] = []
    
    for idx, row in enumerate(all_rows, start=1):
        # Skip if not selected (and selection is active)
        if request.selected_rows is not None and (idx - 1) not in request.selected_rows:
            continue

        # Map columns based on provided mapping
        mapped_row = {}
        for finn_field, source_column in request.column_mapping.dict().items():
            if source_column and source_column in row:
                value = row[source_column]
                # Skip NaN/None values
                if pd.notna(value):
                    mapped_row[finn_field] = value
        
        # Validate row
        row_errors = validate_building_row(mapped_row, idx)
        
        if row_errors:
            errors.append(ImportError(row_number=idx, errors=row_errors))
            if not request.skip_invalid:
                # Fail fast if skip_invalid is False
                return ExecuteImportResponse(
                    success_count=0,
                    skipped_count=len(all_rows),
                    errors=errors,
                    message=f"Tuonti epäonnistui. Korjaa virheet ja yritä uudelleen."
                )
        else:
            mapped_rows.append(mapped_row)
    
    # Get user's org_id from org_users table
    success_count = 0
    db_url = os.environ.get("DATABASE_URL")
    
    if not db_url:
        raise HTTPException(status_code=500, detail="Database configuration missing")
    
    try:
        conn = await asyncpg.connect(db_url)
        
        # Get user's organization
        org_id = await conn.fetchval(
            "SELECT org_id FROM org_users WHERE user_id = $1 LIMIT 1",
            user.sub
        )
        
        if not org_id:
            await conn.close()
            raise HTTPException(
                status_code=403,
                detail="Käyttäjällä ei ole organisaatiota. Ota yhteyttä tukeen."
            )
        
        # Default cost per m2 for Finnish properties (can be updated later by user)
        DEFAULT_COST_PER_M2 = 1800.0
        
        for row in mapped_rows:
            try:
                # Detect sub-building
                is_sub = False
                if request.detect_sub_buildings and row.get('name'):
                    is_sub = is_sub_building(row.get('name'))

                await conn.execute(
                    """
                    INSERT INTO buildings (
                        org_id, name, address, municipality, construction_year, area_m2, 
                        building_type, cost_per_m2, status, is_sub_building
                    ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                    """,
                    org_id,
                    row.get('name'),
                    row.get('address'),
                    row.get('municipality'),
                    int(row['construction_year']),
                    float(row['area_m2']),
                    row.get('building_type'),
                    DEFAULT_COST_PER_M2,
                    'active',
                    is_sub
                )
                success_count += 1
            except Exception as e:
                print(f"Error inserting row: {e}")
                errors.append(ImportError(
                    row_number=mapped_rows.index(row) + 1,
                    errors=[f"Tietokantavirhe: {str(e)}"]
                ))
        
        await conn.close()
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Database error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Tietokantavirhe: {str(e)}"
        )
    
    skipped_count = len(errors)
    message = f"{success_count} rakennusta tuotu onnistuneesti."
    if skipped_count > 0:
        message += f" {skipped_count} riviä ohitettu virheiden vuoksi."
    
    return ExecuteImportResponse(
        success_count=success_count,
        skipped_count=skipped_count,
        errors=errors,
        message=message
    )

@router.post("/pts/execute")
async def execute_pts_import(user: AuthorizedUser, request: ExecutePTSImportRequest) -> ExecuteImportResponse:
    if request.file_id not in UPLOAD_CACHE:
        raise HTTPException(status_code=400, detail="File session expired")
    
    content, filename, header_row = UPLOAD_CACHE[request.file_id]
    
    try:
        filename_lower = filename.lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), header=header_row)
        else:
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=header_row)
        all_rows = df.dropna(how='all').to_dict('records')
        del UPLOAD_CACHE[request.file_id]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Read error: {str(e)}")

    errors = []
    success_count = 0
    conn = await asyncpg.connect(os.environ.get("DATABASE_URL"))
    
    try:
        org_id = await conn.fetchval("SELECT org_id FROM org_users WHERE user_id = $1 LIMIT 1", user.sub)
        if not org_id:
            raise HTTPException(status_code=403, detail="No organization")

        for idx, row in enumerate(all_rows, start=1):
            # Map fields
            data = {}
            for field, col in request.column_mapping.dict().items():
                if col and col in row:
                    data[field] = row[col]

            # Find building
            building_name = data.get('building_name')
            if not building_name:
                errors.append(ImportError(row_number=idx, errors=["Rakennuksen nimi puuttuu"]))
                continue
                
            building_id = await find_building_id(conn, str(building_name), org_id)
            if not building_id:
                errors.append(ImportError(row_number=idx, errors=[f"Rakennusta '{building_name}' ei löydy"]))
                continue

            try:
                cost = parse_float(data.get('cost'))
                year = int(data.get('year')) if data.get('year') and str(data.get('year')).isdigit() else datetime.now().year
                
                # Check for "Funding Source" if it was part of mapping (it's not yet, but optional)
                # Just use description for now
                
                await conn.execute("""
                    INSERT INTO maintenance_tasks (
                        org_id, building_id, title, category, priority, 
                        estimated_cost, status, component_type, description
                    ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                """,
                    org_id,
                    building_id,
                    f"PTS: {data.get('component', 'Muu')} {year}",
                    'renovation',
                    str(data.get('priority', 'medium')).lower(),
                    cost,
                    'planned',
                    str(data.get('component', 'other')).lower(),
                    f"PTS-suunnitelma vuodelle {year}"
                )
                success_count += 1
            except Exception as e:
                errors.append(ImportError(row_number=idx, errors=[str(e)]))

    finally:
        await conn.close()
        
    return ExecuteImportResponse(success_count=success_count, skipped_count=len(errors), errors=errors, message=f"{success_count} PTS-riviä tuotu.")

@router.post("/assessments/execute")
async def execute_assessment_import(user: AuthorizedUser, request: ExecuteAssessmentImportRequest) -> ExecuteImportResponse:
    if request.file_id not in UPLOAD_CACHE:
        raise HTTPException(status_code=400, detail="File session expired")
    
    content, filename, header_row = UPLOAD_CACHE[request.file_id]
    
    try:
        filename_lower = filename.lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), header=header_row)
        else:
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=header_row)
        all_rows = df.dropna(how='all').to_dict('records')
        del UPLOAD_CACHE[request.file_id]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Read error: {str(e)}")

    errors = []
    success_count = 0
    conn = await asyncpg.connect(os.environ.get("DATABASE_URL"))
    
    try:
        org_id = await conn.fetchval("SELECT org_id FROM org_users WHERE user_id = $1 LIMIT 1", user.sub)
        
        # 1. Group rows by building + date to merge component scores
        grouped_assessments = {} # Key: (building_id, date_str), Value: {scores: {}, notes: []}
        
        for idx, row in enumerate(all_rows, start=1):
            data = {}
            for field, col in request.column_mapping.dict().items():
                if col and col in row:
                    data[field] = row[col]

            building_name = data.get('building_name')
            if not building_name:
                errors.append(ImportError(row_number=idx, errors=["Rakennus puuttuu"]))
                continue
                
            building_id = await find_building_id(conn, str(building_name), org_id)
            if not building_id:
                errors.append(ImportError(row_number=idx, errors=[f"Rakennusta '{building_name}' ei löydy"]))
                continue

            # Parse date
            parsed_date = parse_date(data.get('date'))
            if not parsed_date:
                # Fallback to current date or warn? 
                # Let's use current date but log a warning? No, easier to just use current date for now.
                parsed_date = datetime.now().date()
            
            key = (building_id, parsed_date)
            
            if key not in grouped_assessments:
                grouped_assessments[key] = {'scores': {}, 'notes': set()}
            
            # Map component
            component = str(data.get('component', '')).lower()
            try:
                score_raw = data.get('score')
                if not score_raw or pd.isna(score_raw):
                    continue
                score = int(float(str(score_raw).replace(',','.')))
            except:
                continue # Skip invalid scores
                
            # Component mapping
            col_map = {
                'vesikatto': 'facade_roof', 'katto': 'facade_roof', 'julkisivu': 'facade_roof', 'ulkoverhous': 'facade_roof',
                'runko': 'structure', 'perustus': 'structure', 'rakenteet': 'structure', 'sokkeli': 'structure',
                'ikkunat': 'windows_doors', 'ovet': 'windows_doors', 'ikkuna': 'windows_doors', 'ovi': 'windows_doors',
                'lvi': 'plumbing', 'putket': 'plumbing', 'vesi': 'plumbing', 'viemäri': 'plumbing',
                'sähkö': 'electrical', 'sähköt': 'electrical',
                'ilmanvaihto': 'hvac', 'iv': 'hvac',
                'lämmitys': 'heating', 'kaukolämpö': 'heating',
                'sisäpinnat': 'interior_finishes', 'pinnat': 'interior_finishes', 'lattia': 'interior_finishes', 'katto (sisä)': 'interior_finishes',
                'väliseinät': 'interior_walls', 'seinät': 'interior_walls'
            }
            
            db_component = 'structure' # default fallback
            found = False
            for k, v in col_map.items():
                if k in component:
                    db_component = v
                    found = True
                    break
            
            grouped_assessments[key]['scores'][f"{db_component}_score"] = score
            grouped_assessments[key]['notes'].add(f"{component}: {score}")

        # 2. Insert grouped assessments
        for (building_id, a_date), info in grouped_assessments.items():
            try:
                scores = info['scores']
                notes = "; ".join(info['notes'])
                
                # Check if assessment already exists for this date
                existing_id = await conn.fetchval(
                    "SELECT id FROM component_assessments WHERE building_id = $1 AND assessment_date = $2",
                    building_id, a_date
                )
                
                if existing_id:
                    # Update existing
                    set_clauses = []
                    values = [existing_id]
                    for i, (col, val) in enumerate(scores.items(), start=2):
                        set_clauses.append(f"{col} = ${i}")
                        values.append(val)
                    
                    if set_clauses:
                        await conn.execute(
                            f"UPDATE component_assessments SET {', '.join(set_clauses)}, notes = notes || '; ' || '{notes}' WHERE id = $1",
                            *values
                        )
                else:
                    # Insert new
                    cols = ['building_id', 'assessment_date', 'notes'] + list(scores.keys())
                    placeholders = [f"${i+1}" for i in range(len(cols))]
                    vals = [building_id, a_date, notes] + list(scores.values())
                    
                    await conn.execute(
                        f"INSERT INTO component_assessments ({', '.join(cols)}) VALUES ({', '.join(placeholders)})",
                        *vals
                    )
                
                success_count += 1
            except Exception as e:
                print(f"Error saving assessment: {e}")
                errors.append(ImportError(row_number=0, errors=[f"Tallennusvirhe {building_id}: {str(e)}"]))

    finally:
        await conn.close()

    return ExecuteImportResponse(success_count=success_count, skipped_count=len(errors), errors=errors, message=f"{success_count} kuntoarviota käsitelty.")

@router.post("/maintenance/execute")
async def execute_maintenance_import(user: AuthorizedUser, request: ExecuteMaintenanceImportRequest) -> ExecuteImportResponse:
    # Maintenance history import
    if request.file_id not in UPLOAD_CACHE:
        raise HTTPException(status_code=400, detail="File session expired")
    
    content, filename, header_row = UPLOAD_CACHE[request.file_id]
    
    try:
        filename_lower = filename.lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content), header=header_row)
        else:
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=header_row)
        all_rows = df.dropna(how='all').to_dict('records')
        del UPLOAD_CACHE[request.file_id]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Read error: {str(e)}")

    errors = []
    success_count = 0
    conn = await asyncpg.connect(os.environ.get("DATABASE_URL"))
    
    try:
        org_id = await conn.fetchval("SELECT org_id FROM org_users WHERE user_id = $1 LIMIT 1", user.sub)
        
        for idx, row in enumerate(all_rows, start=1):
            data = {}
            for field, col in request.column_mapping.dict().items():
                if col and col in row:
                    data[field] = row[col]

            building_name = data.get('building_name')
            if not building_name:
                errors.append(ImportError(row_number=idx, errors=["Rakennus puuttuu"]))
                continue
                
            building_id = await find_building_id(conn, str(building_name), org_id)
            if not building_id:
                errors.append(ImportError(row_number=idx, errors=[f"Rakennusta '{building_name}' ei löydy"]))
                continue

            try:
                cost = parse_float(data.get('cost'))
                completed_date = parse_date(data.get('date'))
                if not completed_date:
                    completed_date = datetime.now().date() # Fallback? Or skip? Let's fallback.
                
                await conn.execute("""
                    INSERT INTO maintenance_tasks (
                        org_id, building_id, title, category, priority, 
                        actual_cost, status, component_type, description, completed_date
                    ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                """,
                    org_id,
                    building_id,
                    str(data.get('description', 'Huolto')),
                    'repair',
                    'medium',
                    cost,
                    'completed',
                    str(data.get('component', 'other')).lower(),
                    f"Historiatieto: {data.get('description')}",
                    completed_date
                )
                success_count += 1
            except Exception as e:
                errors.append(ImportError(row_number=idx, errors=[str(e)]))

    finally:
        await conn.close()

    return ExecuteImportResponse(success_count=success_count, skipped_count=len(errors), errors=errors, message=f"{success_count} huoltotietoa tuotu.")
